from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Prijsvergelijking"

# ── Data ──────────────────────────────────────────────────────────────────────
concurrenten = [
    "plissehordeurenwebshop.nl",
    "unilux.nl",
    "horrentotaal.nl",
    "horrengigant.nl",
    "horren.com",
    "praxis.nl",
    "luxaflex.nl",
    "creon-kozijnen.nl",
    "gamma.nl",
    "bruynzeelhomeproducts.nl",
]

producten = [
    "Enkele deur – Klein",
    "Enkele deur – Middel",
    "Enkele deur – Groot",
    "Dubbele deur – Klein",
    "Dubbele deur – Middel",
    "Dubbele deur – Groot",
]

afmetingen = {
    "Enkele deur – Klein":   "730 × 1970 mm",
    "Enkele deur – Middel":  "870 × 2120 mm",
    "Enkele deur – Groot":   "1030 × 2270 mm",
    "Dubbele deur – Klein":  "1430 × 1970 mm",
    "Dubbele deur – Middel": "1630 × 2120 mm",
    "Dubbele deur – Groot":  "1830 × 2270 mm",
}

# Known prices / statuses (to be filled in after running Playwright tests)
# "?" = niet uitgelezen, "Op aanvraag" = geen online prijs
known = {
    "luxaflex.nl":            {p: "Op aanvraag" for p in producten},
    "bruynzeelhomeproducts.nl": {p: "Op aanvraag" for p in producten},
}

# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREY_BG    = "F2F2F2"
WHITE      = "FFFFFF"
ORANGE     = "FF6600"
GREEN_TXT  = "1E7145"

thin = Side(border_style="thin", color="BFBFBF")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

header_font       = Font(name="Arial", bold=True, color=WHITE, size=10)
subheader_font    = Font(name="Arial", bold=True, color=WHITE, size=9)
col_header_font   = Font(name="Arial", bold=True, color=WHITE, size=9)
product_font      = Font(name="Arial", bold=True, size=9)
size_font         = Font(name="Arial", italic=True, color="595959", size=8)
cell_font         = Font(name="Arial", size=9)
na_font           = Font(name="Arial", italic=True, color="A6A6A6", size=9)
own_col_font      = Font(name="Arial", bold=True, color=ORANGE, size=9)
own_header_font   = Font(name="Arial", bold=True, color=WHITE, size=9)

fill_dark     = PatternFill("solid", fgColor=DARK_BLUE)
fill_mid      = PatternFill("solid", fgColor=MID_BLUE)
fill_light    = PatternFill("solid", fgColor=LIGHT_BLUE)
fill_grey     = PatternFill("solid", fgColor=GREY_BG)
fill_orange   = PatternFill("solid", fgColor=ORANGE)
fill_white    = PatternFill("solid", fgColor=WHITE)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Layout ────────────────────────────────────────────────────────────────────
# Row 1 : title (merged)
# Row 2 : subtitle
# Row 3 : empty spacer
# Row 4 : column headers (Product | Afmeting | concurrent…)
# Row 5+ : data

TITLE_ROW   = 1
SUBTTL_ROW  = 2
HEADER_ROW  = 4
DATA_START  = 5

COL_PRODUCT = 1   # A
COL_SIZE    = 2   # B
COL_FIRST   = 3   # C  → first competitor

n_cols = COL_FIRST + len(concurrenten) - 1  # last data column

# ── Title ─────────────────────────────────────────────────────────────────────
ws.merge_cells(start_row=TITLE_ROW, start_column=1,
               end_row=TITLE_ROW, end_column=n_cols)
title_cell = ws.cell(TITLE_ROW, 1,
    "Prijsvergelijking – Plissé Hordeuren Op Maat")
title_cell.font      = Font(name="Arial", bold=True, color=WHITE, size=14)
title_cell.fill      = fill_dark
title_cell.alignment = center

ws.merge_cells(start_row=SUBTTL_ROW, start_column=1,
               end_row=SUBTTL_ROW, end_column=n_cols)
sub = ws.cell(SUBTTL_ROW, 1,
    "Configuratie: RAL 9010 wit • Zwart gaas • In de dag montage  |  "
    "Maten zijn NIET-standaard (maatwerk)  |  Prijzen incl. BTW")
sub.font      = Font(name="Arial", italic=True, color="D9D9D9", size=9)
sub.fill      = fill_mid
sub.alignment = center

# ── Column headers ────────────────────────────────────────────────────────────
ws.cell(HEADER_ROW, COL_PRODUCT, "Product").font      = subheader_font
ws.cell(HEADER_ROW, COL_PRODUCT).fill      = fill_mid
ws.cell(HEADER_ROW, COL_PRODUCT).alignment = center
ws.cell(HEADER_ROW, COL_PRODUCT).border    = border_thin

ws.cell(HEADER_ROW, COL_SIZE, "Afmeting (B × H)").font      = subheader_font
ws.cell(HEADER_ROW, COL_SIZE).fill      = fill_mid
ws.cell(HEADER_ROW, COL_SIZE).alignment = center
ws.cell(HEADER_ROW, COL_SIZE).border    = border_thin

for idx, naam in enumerate(concurrenten):
    col = COL_FIRST + idx
    c = ws.cell(HEADER_ROW, col, naam)
    if naam == "plissehordeurenwebshop.nl":
        c.font = own_header_font
        c.fill = fill_orange
    else:
        c.font = col_header_font
        c.fill = fill_mid
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, text_rotation=0)
    c.border = border_thin

# ── Data rows ─────────────────────────────────────────────────────────────────
for r_idx, product in enumerate(producten):
    row = DATA_START + r_idx
    bg = fill_light if r_idx % 2 == 0 else fill_white

    # Product name
    pc = ws.cell(row, COL_PRODUCT, product)
    pc.font      = product_font
    pc.fill      = bg
    pc.alignment = left
    pc.border    = border_thin

    # Size
    sc = ws.cell(row, COL_SIZE, afmetingen[product])
    sc.font      = size_font
    sc.fill      = bg
    sc.alignment = center
    sc.border    = border_thin

    # Competitor prices
    for idx, naam in enumerate(concurrenten):
        col = COL_FIRST + idx
        val = known.get(naam, {}).get(product, "?")
        c = ws.cell(row, col, val)
        c.alignment = center
        c.border    = border_thin
        if val == "Op aanvraag":
            c.font = na_font
            c.fill = PatternFill("solid", fgColor="F2F2F2")
        elif val == "?":
            c.font = Font(name="Arial", color="A6A6A6", size=9)
            c.fill = bg
        elif naam == "plissehordeurenwebshop.nl":
            c.font = Font(name="Arial", bold=True, color=ORANGE, size=9)
            c.fill = PatternFill("solid", fgColor="FFF2E5")
        else:
            c.font = cell_font
            c.fill = bg

# ── Notes row ────────────────────────────────────────────────────────────────
note_row = DATA_START + len(producten) + 1
ws.merge_cells(start_row=note_row, start_column=1,
               end_row=note_row, end_column=n_cols)
note = ws.cell(note_row, 1,
    "Legenda:  ? = prijs nog niet uitgelezen (voer Playwright tests uit)  │  "
    "Op aanvraag = geen online prijs beschikbaar (dealer/showroom model)  │  "
    "Oranje = eigen webshop (plissehordeurenwebshop.nl)")
note.font      = Font(name="Arial", italic=True, color="595959", size=8)
note.fill      = PatternFill("solid", fgColor="FFFBE6")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
note.border    = border_thin

# ── Playwright-instructie blad ───────────────────────────────────────────────
ws2 = wb.create_sheet("Playwright Instructies")
instructions = [
    ["Stap", "Actie"],
    ["1", "Open terminal in de map playwright-prijzen/"],
    ["2", "Voer uit: npm install"],
    ["3", "Voer uit: npx playwright install chromium"],
    ["4", "Voer uit: npm test   (draait alle 10 tests)"],
    ["5", "Of voer één test uit, bijv.: npm run test:eigen"],
    ["6", "Resultaten staan in results/prices.json"],
    ["7", "Kopieer de prijzen uit de console-output naar dit Excel-bestand"],
    ["", ""],
    ["Test bestand", "Concurrent"],
    ["01-plissehordeurenwebshop.spec.js", "plissehordeurenwebshop.nl (eigen)"],
    ["02-unilux.spec.js", "unilux.nl"],
    ["03-horrentotaal.spec.js", "horrentotaal.nl"],
    ["04-horrengigant.spec.js", "horrengigant.nl"],
    ["05-horren-com.spec.js", "horren.com"],
    ["06-praxis.spec.js", "praxis.nl"],
    ["07-luxaflex.spec.js", "luxaflex.nl (Op aanvraag – dealer)"],
    ["08-creon-kozijnen.spec.js", "creon-kozijnen.nl"],
    ["09-gamma.spec.js", "gamma.nl"],
    ["10-bruynzeel.spec.js", "bruynzeelhomeproducts.nl (Op aanvraag – dealer)"],
]

ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 45

for r, row_data in enumerate(instructions, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(r, c, val)
        cell.font      = Font(name="Arial", bold=(r in (1, 10)), size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if r == 1 or r == 10:
            cell.fill = fill_mid
            cell.font = Font(name="Arial", bold=True, color=WHITE, size=9)

# ── Column widths (main sheet) ────────────────────────────────────────────────
ws.column_dimensions["A"].width = 24  # Product
ws.column_dimensions["B"].width = 18  # Afmeting
for idx in range(len(concurrenten)):
    col_letter = get_column_letter(COL_FIRST + idx)
    ws.column_dimensions[col_letter].width = 22

ws.row_dimensions[TITLE_ROW].height  = 28
ws.row_dimensions[SUBTTL_ROW].height = 18
ws.row_dimensions[HEADER_ROW].height = 36
for r in range(DATA_START, DATA_START + len(producten)):
    ws.row_dimensions[r].height = 22
ws.row_dimensions[note_row].height = 28

# Freeze panes: lock rows 1-4 and columns A-B
ws.freeze_panes = "C5"

out = "/sessions/stoic-upbeat-babbage/mnt/outputs/playwright-prijzen/prijsvergelijking-plisse-hordeuren.xlsx"
wb.save(out)
print("Saved:", out)
